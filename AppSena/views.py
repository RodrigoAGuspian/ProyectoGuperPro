from django.shortcuts import render, redirect
from .models import *
from .forms import * 
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from datetime import datetime, date
from openpyxl import load_workbook #para cargar archivos con esxtencion xl


from itertools import chain #para combinar 2 o mas querysets
from operator import attrgetter
#=======================INDEX=====================#
def view_index(request):
	if request.user.is_authenticated:
		return render(request, 'index.html')
	else:
		return redirect('url_login')
#=================================================#

#=======================FICHA=====================#
def view_lista_fichas(request):
	ficha = Ficha.objects.all()
	if not ficha:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request, 'lista_fichas.html', locals())
#=================================================#

#=================AGREGAR FICHA===================#
def view_agregar_ficha(request):

	msj= 'Agregar Ficha'

	if request.method == "POST":
		formulario_ficha = agregar_ficha_form(request.POST)
		if formulario_ficha.is_valid():

			h = date.today()
			fic = formulario_ficha.save(commit = False)

			fecha = formulario_ficha.cleaned_data['fechaFinEtapaLectiva']
			ambte = formulario_ficha.cleaned_data['ambiente']
			jrnad = formulario_ficha.cleaned_data['jornada']

			dispone = Ficha.objects.filter(ambiente=ambte, jornada=jrnad)

			if dispone:
				for i in dispone:
					ocupado = str(i.numeroFicha)+' de '+str(i.programa.nombre)
					msj = 'A ocurrido un error al intentar Agregar la Ficha'
					msj2='El anbiete seleccionado esta siendo ocupado en el mismo horario por la ficha numero: '+ocupado

			elif fecha>h:
				fic.save()
				return redirect('url_lista_fichas')
			else:
				msj = 'A ocurrido un error al intentar Agregar la Ficha'
				msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'

	else:
		formulario_ficha = agregar_ficha_form()
	return render(request, 'agregar_ficha.html', locals())
#=================================================#

#================AGREGAR PERSONA FICHA============#
def view_agregar_persona_ficha(request):
	if request.method == "POST":
		formulario_persona_ficha = agregar_persona_ficha_form(request.POST)
		if formulario_persona_ficha.is_valid():
			persona_ficha = formulario_persona_ficha.save(commit = False)
			persona_ficha.save()
			formulario_persona_ficha.save()

			return redirect('url_lista_fichas')
	else:
		formulario_persona_ficha = agregar_persona_ficha_form()
	return render(request, 'agregar_persona_ficha.html', locals())
#=================================================#

#===================EDITAR FICHA==================#
def view_editar_ficha(request, id_ficha):

	msj='Editar Ficha'

	ficha = Ficha.objects.get(id = id_ficha)
	if request.method == "POST":
		formulario_ficha = agregar_ficha_form(request.POST, instance = ficha)
		if formulario_ficha.is_valid():

			h = date.today()

			fecha = formulario_ficha.cleaned_data['fechaFinEtapaLectiva']
			ambte = formulario_ficha.cleaned_data['ambiente']
			jrnad = formulario_ficha.cleaned_data['jornada']

			dispone = Ficha.objects.filter(ambiente=ambte, jornada=jrnad)
			
			if not dispone: #si el Query esta vacio el ambiente esta disponible
				
				if fecha>h:
					formulario_ficha.save()
					return redirect('url_lista_fichas')
				else:
					msj = 'A ocurrido un error al intentar Editar la Ficha'
					msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'	
			else:
				for i in dispone:
					#VALORES SIN APLICAR CABIOS
					idfichDis =i.id  		       #id de la ficha SIN cambios
					numfichDis =i.numeroFicha      #numero de la ficha SIN cambios
					profichDis = i.programa.nombre #programa de la ficha SIN cambios
					#VALORES CON LOS CABIOS QUE SE QUIEREN APLICAR
					ficid = ficha.id    		   #id de la ficha CON cambios			
					if idfichDis==ficid:
						
						if fecha>h:
							formulario_ficha.save()
							return redirect('url_lista_fichas')
						else:
							msj = 'A ocurrido un erro al intentar Editar la Ficha'
							msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'
					else:
						ocupado = numfichDis+' de '+profichDis
						msj = 'A ocurrido un erro al intentar Editar la Ficha'
						msj2='El anbiete esta siendo ocupado en el mismo horario por la ficha numero: '+ocupado
						
	else:
		formulario_ficha = agregar_ficha_form(instance = ficha)
	return render(request, 'agregar_ficha.html', locals())
#=================================================#

#==================BORRAR FICHA===================#
def view_eliminar_ficha(request, id_ficha):
	try:
		fc = Persona_ficha.objects.filter(ficha=id_ficha)
		if fc:
			msj = ' ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas'
			print(">>>>  ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas")
		else:
			ficha = Ficha.objects.get(id = id_ficha)
			ficha.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_fichas')
#=================================================#

#=====================PERMISO=====================#
def view_lista_permisos(request):
	permiso = Permiso.objects.all()
	return render(request, 'lista_permisos.html', locals())
#=================================================#

#=====================PERSONA=====================#

#========== REGISTRAR ADMINISTRADOR ==============#
def view_agregar_administrador(request):
	
	msj = 'Agregar Administrador'

	r = Rol.objects.filter(rol='ADMINISTRADOR')
	if not r:
		msjErrorRol = 'Debes agregar el rol de ADMINISTRADOR'

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		formulario4 = elegir_rol_admin_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			rol = formulario4.cleaned_data['rol']

			try:
				r = Rol.objects.get(rol=rol)
				query = User.objects.filter(email=email)
				if not query:
					try:
						u = User.objects.create_user(username=username,password=documento, email=email)
						persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
						admin = Rol_persona.objects.create(rol=r, persona=persona)
						
						u.save()
						persona.save()
						admin.save()

						return redirect ('url_agregar_admin')
					except:
						msjError = '¡El CORREO ya existe!'
				else:
					msjError = 'Este CORREO ya esta registrado'
			except:
				msjErrorRol = '¡El Rol ADMINISTRADOR aun no esta registrado'
			#=======
			
	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()
		formulario4 = elegir_rol_admin_form()

	return render(request, 'agregar_administrador.html', locals())
#=================================================#

#==============LISTA INSTRUCTORES=================#
def view_lista_instructores(request):

	try:
		r = Rol.objects.get(rol='INSTRUCTOR')
		rp = Rol_persona.objects.filter(rol=r)
		
		if not rp:
			msj = 'Lista Vacia. Click aqui para agregar'

	except:
		msj2 = 'Debes agregar El rol INSTRUCTOR'

	return render (request, 'lista_instructores.html', locals())
#=================================================#

#=================REGISTRAR INSTRUCTOR============#
def view_agregar_instructor(request):

	msj = 'Agregar Instructor'
	
	r = Rol.objects.filter(rol='INSTRUCTOR')

	if not r:
		msjErrorRol = '¡El Rol INSTRUCTOR aun no esta registrado'

	if request.method == "POST":
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		formulario4 = elegir_rol_instructor_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			rol = formulario4.cleaned_data['rol']

			try:
				r = Rol.objects.get(rol=rol)
				query = User.objects.filter(email=email)
				if not query:
					try:
						u = User.objects.create_user(username=username,password=documento, email=email)
						persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
						vigilante = Rol_persona.objects.create(rol=r, persona=persona)
						
						u.save()
						persona.save()
						vigilante.save()

						return redirect ('url_lista_instructores')
					except:
						msjError = '¡El CORREO ya existe!'
				else:
					msjError = 'Este CORREO ya esta registrado'
			except:
				msjErrorRol = '¡El Rol INSTRUCTOR aun no esta registrado'

	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()
		formulario4 = elegir_rol_instructor_form()

	return render(request, 'agregar_instructor.html', locals())
#=================================================#

#==============EDITAR_INSTRUCTOR==================#
def view_editar_instructor(request, id_instructor):

	msj = 'Editar Instructor'
	ins = Persona.objects.get(id=id_instructor)
	#fic = Persona_ficha.objects.get(persona=ins)
	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES, instance=ins)
		formulario2 = editar_user_form(request.POST, request.FILES, instance=ins.usuario)
		#formulario3 = elegir_ficha_form(request.POST, request.FILES, instance=fic)
		if formulario.is_valid() and formulario2.is_valid():# and formulario3.is_valid():

			try:
				query = User.objects.get(email = ins.usuario.email)
				query_id = query.id

				if  query_id == ins.usuario.id:
					formulario.save()
					formulario2.save()
					return redirect ('url_lista_instructores')

				else:
					msjError = 'El correo que quieres ingresar ya existe'
			except:
				formulario.save()
				formulario2.save()
				return redirect ('url_lista_instructores')

	else:
		formulario = agregar_persona_form(instance=ins)
		formulario2 = editar_user_form(instance=ins.usuario)
		#formulario3 = elegir_ficha_form(instance=fic)
	return render (request, 'agregar_instructor.html', locals())
#=================================================#

#==============ELIMINAR_INSTRUCTOR================#
def view_eliminar_instructor(request, id_instructor):
	try:
		instructor = Persona.objects.get(id=id_instructor)
		actividad = Permiso_persona.objects.filter(instructor=id_instructor)
		
		if actividad:
			msj2="Upps a ocurrido un inconveniente. el VIGILANTE tiene actividad"
			print(msj2)
		else:
			id_user = instructor.usuario.id
			user = User.objects.get(id=id_user)
			instructor.delete()
			user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect ('url_lista_instructores')
#=================================================#

#==============LISTA_VIGILANTES==================#
def view_lista_vigilantes(request):
	try:
		vg= Rol.objects.get(rol='VIGILANTE')
		vig=Rol_persona.objects.filter(rol=vg)
		if not vig:
			msj = 'Lista Vacia. Click aqui para agregar'
	except:
		msj2 = 'Debes agregar El rol Vigilante'

	return render(request, 'lista_vigilantes.html', locals())
#================================================#

#===========REGISTRAR VIGILANTE===================#
def view_agregar_vigilante(request):

	msj = 'Agregar Vigilante'

	r = Rol.objects.filter(rol='VIGILANTE')

	if not r:
		msjErrorRol = '¡El Rol VIGILANTE aun no esta registrado'

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_vigilante_form(request.POST, request.FILES)
		formulario4 = elegir_rol_vigilante_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']

			rol = formulario4.cleaned_data['rol']

			
			try:
				r = Rol.objects.get(rol=rol)
				query = User.objects.filter(email=email)
				print('>>>>>>>>>>>>>>>>  ',query)
				if not query:
					try:
						u = User.objects.create_user(username=email,password=documento, email=email)
						persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
						vigilante = Rol_persona.objects.create(persona=persona, rol=r)
						
						u.save()
						persona.save()
						vigilante.save()
						
						return redirect ('url_lista_vigilantes')
					except:
						msjError = '¡El CORREO ya existe!'
				else:
					msjError = 'Este CORREO ya esta registrado'
			except:
				msjErrorRol = '¡El Rol VIGILANTE aun no esta registrado'

	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_vigilante_form()
		formulario4 = elegir_rol_vigilante_form()
	return render(request, 'agregar_vigilante.html', locals())	
#=================================================#

#==============EDITAR_VIGILANTE===================#
def view_editar_vigilante(request, id_vigilante):

	msj = 'Editar Vigilante'

	vgt= Persona.objects.get(id= id_vigilante)
	docOld = vgt.documentoIdentidad
	if request.method== 'POST':
		formulario= agregar_persona_form(request.POST, request.FILES, instance= vgt)
		formulario2= editar_user_form(request.POST, request.FILES, instance= vgt.usuario)
		if formulario.is_valid() and formulario2.is_valid():

			doc = formulario.cleaned_data['documentoIdentidad']

			try:
				query = User.objects.get(email = vgt.usuario.email)
				query_id = query.id

				if  query_id == vgt.usuario.id:
					if doc!=docOld:
						vgt.usuario.set_password(doc)

					formulario.save()
					formulario2.save()
					return redirect('url_lista_vigilantes')

				else:
					msjError = 'El correo que quieres ingresar ya existe'
			except:
				if doc!=docOld:
						vgt.usuario.set_password(doc)
				formulario.save()
				formulario2.save()
				return redirect('url_lista_vigilantes')
			
	else:
		formulario = agregar_persona_form(instance=vgt)
		formulario2 = editar_user_form(instance=vgt.usuario)
	return render (request, 'agregar_instructor.html', locals())
#=================================================#

#==============ELIMINAR_VIGILANTE=================#
def view_eliminar_vigilante(request, id_vigilante):
	try:
		vigilante = Persona.objects.get(id=id_vigilante)
		actividad = Permiso_persona.objects.filter(vigilante=id_vigilante)

		if actividad:
			msj2="Upps a ocurrido un inconveniente. el VIGILANTE tiene actividad"
			print(msj2)
		else:
			id_user = vigilante.usuario.id
			user = User.objects.get(id=id_user)
			vigilante.delete()
			user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"
		print(msj2)
	return redirect ('url_lista_vigilantes')
#=================================================#

#==============LISTA_APRENDICES===================#
def view_lista_aprendices(request):

	try:
		r = Rol.objects.get(rol='APRENDIZ')
		rp = Rol_persona.objects.filter(rol=r)
		
		if not rp:
			msj = 'Lista Vacia. Click aqui para agregar'

	except:
		msj2 = 'Debes agregar El rol Aprendiz'

	return render (request, 'lista_aprendices.html', locals())
#=================================================#

#==============AGREGAR_APRENDIZ===================#
def view_agregar_aprendiz(request):

	msj = 'Agregar Aprendiz'

	r = Rol.objects.filter(rol='APRENDIZ')
	print(r)

	if not r:
		msjErrorRol = '¡El Rol APRENDIZ aun no esta registrado'

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2 = agregar_user_form(request.POST, request.FILES)
		formulario3 = elegir_ficha_form(request.POST, request.FILES)
		formulario4 = elegir_rol_aprendiz_form(request.POST, request.FILES)
		
		if formulario.is_valid() and formulario2.is_valid() and formulario3.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']
	
			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			fic = formulario3.cleaned_data['ficha']
			rol = formulario4.cleaned_data['rol']

			#Objetos
			try:
				r = Rol.objects.get(rol=rol)

				query = User.objects.filter(email=email)
				if not query:
					try:
						u = User.objects.create_user(username=username,password=documento, email=email)
						persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
						fich_per = Persona_ficha.objects.create(persona=persona, ficha=fic)
						aprendiz = Rol_persona.objects.create(persona=persona, rol=r)
						#Guardar Objetos
						u.save()
						persona.save()			
						fich_per.save()
						aprendiz.save()

						return redirect ('url_lista_aprendices')
					except:
						msjError = '¡Este CORREO ya esta registrado'
				else:
					msjError = 'Este CORREO ya esta registrado'
			except:
				msjErrorRol = '¡El Rol APRENDIZ aun no esta registrado'

		else:
			print('FORMULARIO INVALIDO')

	else:
		formulario = agregar_persona_form()
		formulario2 = agregar_user_form()
		formulario3 = elegir_ficha_form()
		formulario4 = elegir_rol_aprendiz_form()

	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#

#==============EDITAR_APRENDIZ====================#
def view_editar_aprendiz(request, id_aprendiz):

	msj = 'Editar APRENDIZ'

	apr = Persona.objects.get(id=id_aprendiz)
	fic = Persona_ficha.objects.get(persona=apr)
	docOld = apr.documentoIdentidad

	if request.method == "POST":
		formulario  = agregar_persona_form(request.POST, request.FILES, instance=apr)
		formulario2 = editar_user_form(request.POST, request.FILES, instance=apr.usuario)
		formulario3 = elegir_ficha_form(request.POST, request.FILES, instance=fic)
		if formulario.is_valid() and formulario2.is_valid() and formulario3.is_valid():
				
			doc = formulario.cleaned_data['documentoIdentidad']	

			try:
				query = User.objects.get(email = apr.usuario.email)
				query_id = query.id				

				if  query_id == apr.usuario.id:

					if doc!=docOld:
						apr.usuario.set_password(doc)

					formulario.save()
					formulario2.save()
					formulario3.save()
					return redirect ('url_lista_aprendices')

				else:
					msjError = 'El correo que quieres ingresar ya existe'
			except:
				if doc!=docOld:
					apr.usuario.set_password(doc)

				formulario.save()
				formulario2.save()
				formulario3.save()
				return redirect ('url_lista_aprendices')


	else:
		formulario = agregar_persona_form(instance=apr)
		formulario2 = editar_user_form(instance=apr.usuario)
		formulario3 = elegir_ficha_form(instance=fic)
	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#


#==============ELIMINAR_APRENDIZ==================#
def view_eliminar_aprendiz(request, id_aprendiz):
	try:

		actividad = Permiso_persona.objects.filter(persona=id_aprendiz)
		if actividad:
			msj2="Upps a ocurrido un inconveniente. el aprendiz tiene actividad"
			print(msj2)
		else:
			aprendiz = Persona.objects.get(id=id_aprendiz)
			id_user = aprendiz.usuario.id
			user = User.objects.get(id=id_user)
			aprendiz.delete()
			user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect ('url_lista_aprendices')
#=================================================#

#================SUBIR_ARCHIVO_EXCEL==============#
def subir_archivo_excel(f):
	with open('media/registros_excel/registros.xlsx', 'wb+') as destination:
		for chunk in f.chunks():
			destination.write(chunk)
#=================================================#			

#===============VIEW_BUSTCAR_EXCEL================#
def view_agregar_aprendiz_excel(request):

	msj = 'Subir Registros desde Excel'
	try:
		rl = Rol.objects.get(rol='APRENDIZ')

		if request.method == "POST":
			form = cargar_excel_form(request.POST, request.FILES)
			if form.is_valid():
				
				try:
					subir_archivo_excel(request.FILES['docfile'])
					list_xl = cargar_excel()
					
					if not list_xl:
						return redirect('url_lista_aprendices')
					else:
						msjError = 'No se pueden registrar los siguienetes aprendices'

				except:
					msjError = 'Error al Intenatr subir este elemento'
				
			else:
				print('>>>>>>>>>>>>No Valido')
		else:
			form = cargar_excel_form()
	except:
		msjErrorRol = 'El rol APRENDIZ aun no esta registrado'

	return render(request, 'agregar_aprendices_excel.html', locals())
#=================================================#

#====================CARGAR_EXCEL=================#
def cargar_excel():

	documento = ''
	nom	 = ''
	ape = ''
	email = ''
	fic = ''
	tel = ''
	
	lista = []

	rl = Rol.objects.get(rol='APRENDIZ')
	FILE_PATH = 'media/registros_excel/registros.xlsx' #define el achivo excel en una variable
	#hoja=wb.get_sheet_by_name(str(wb.get_sheet_names()[0])) #obtiene el nombre de la hoja
	SHEET = 'Hoja1'                  #define la hoja de excel en una variable

	wb = load_workbook(FILE_PATH, read_only=True)# cargo el archivo en una variable
	sheet = wb[SHEET]                            # cargo la hoja en una variable

	for row in sheet.iter_rows(min_row=2):

		documento = row[0].value
		nom = row[1].value
		ape = row[2].value
		email = row[3].value
		fic = row[5].value
		tel = row[6].value

		try:

			fc = str(fic)
			ficha = Ficha.objects.get(numeroFicha=fc)
			if 'misena.edu.co' in email:

				try:
					query = User.objects.get(email=email)
					p = 'Ya existe un Email igual ',documento,nom,ape,email,fic
					lista.append(p)
				except:
					try:
						u = User.objects.create_user(username=email,password=documento, email=email)		
						try:
							persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
							aprendiz = Rol_persona.objects.create(rol=rl ,persona=persona)
							fich_per = Persona_ficha.objects.create(persona=persona, ficha=ficha)
							
							u.save()
							persona.save()
							aprendiz.save()
							fich_per.save()
						except:
							p = 'Ya existe un numero de identidad igual ',documento,nom,ape,email,fic
							lista.append(p)
					
					except:
						p = 'Ya existe un Email igual ',documento,nom,ape,email,fic
						lista.append(p)

			else:
				p = 'Email no es un correo SENA ',documento,nom,ape,email,fic
				lista.append(p)
		except:
			p = 'La Ficha No Exsiste ',documento,nom,ape,email,fic
			lista.append(p)

	return lista
#=================================================#

#====================PROGRAMA=====================#
def view_lista_programas(request):
	programa = Programa.objects.all()
	if not programa:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request, 'lista_programas.html', locals())
#=================================================#

#===============AGREGAR PROGRAMA==================#
def view_agregar_programa(request):
	msj='Agregar Programa'
	if request.method == 'POST':
		formulario = agregar_programa_form(request.POST, request.FILES)
		if formulario.is_valid():
			pro = formulario.save(commit=False)
			pro.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form()

	return render (request, 'agregar_programa.html',locals())
#=================================================#

#===================EDITAR PROGRAMA===============#	
def view_editar_programa(request, id_programa):
	msj='Editar Programa'
	pro = Programa.objects.get(id=id_programa)
	if request.method == "POST":
		formulario = agregar_programa_form(request.POST, request.FILES, instance=pro)
		if formulario.is_valid():
			pro = formulario.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form(instance= pro)
	return render(request, 'agregar_programa.html', locals())
#=================================================#

#===================ELIMINAR PROGRAMA=============#
def view_eliminar_programa(request, id_programa):
	try:	
		pr = Ficha.objects.filter(programa=id_programa)
		if pr:
			msj = ' ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas'
		else:
			pro = Programa.objects.get(id=id_programa)
			pro.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_programas')	
#=================================================#

#=====================USUARIO=====================#
def view_usuario(request):
	usuario = User.objects.all()
	return render(request, 'usuario/index.html', locals())
#=================================================#

#===================VER_ROL=======================#
def view_lista_roles(request):
	rol = Rol.objects.all()
	if not rol:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request,'lista_roles.html',locals())
#=================================================#

#===================AGREGAR ROL===================#
def view_agregar_rol(request):
	msj='Agregar Rol'
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST,request.FILES)
		if formulario.is_valid():
			rol= formulario.save(commit= False)
			rol.save()
			formulario.save()
			return redirect('url_lista_roles')

	else:
		formulario = agregar_rol_form()
	return render(request,'agregar_rol.html',locals())
#=================================================#

#============ AGREGAR ROL AUTOMATIC================#
def view_agregar_rol_automatic(request, rol):
	try:
		rl = Rol.objects.create(rol=rol)
		rl.save()

		if rol=='APRENDIZ':
			return redirect('url_agregar_aprendiz')
		elif rol=='XL':
			return redirect('url_agregar_aprendiz_excel')
		elif rol=='INSTRUCTOR':
			return redirect('url_agregar_instructor')
		elif rol=='VIGILANTE':
			return redirect('url_agregar_vigilante')
		else:
			return redirect('url_agregar_admin')
	except:
		print('Error Desconocido, vuelve a intentarlo')
#=================================================#

#===================EDITAR_ROL====================#
def view_editar_rol(request, id_rol):
	msj='Editar Rol'
	rol= Rol.objects.get(id=id_rol)
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST, request.FILES, instance=rol)
		if formulario.is_valid():
			rol= formulario.save()
			return redirect('url_lista_roles')
	else:
		formulario= agregar_rol_form(instance= rol)
	return render(request, 'agregar_rol.html',locals())
#=================================================#

#===================ELIMINAR_ROL==================#
def view_eliminar_rol(request, id_rol):
	try:
		rp = Rol_persona.objects.filter(rol=id_rol)
		if rp:
			msj = ' ¡Inposible Borrar! Este Rol tiene personas asociadas'
			print(">>>> "+msj)
		else:
			rol= Rol.objects.get(id= id_rol)
			rol.delete()
	except:
		msj="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_roles')
#==================== LOGIN ======================#
#@login_required  DECORADORES
def view_login(request):
	error = "Usted no esta autorizado para ingresar desde este terminal. Por favor dirijase a su Teléfono."
	comprobacion_rol = ""
	info_enviada = False
	if request.user.is_authenticated:
		return redirect('url_index')
	else:
		loginf = login_form()
		if request.method == "POST":
			loginf = login_form(request.POST)
			if loginf.is_valid():
				usuario = loginf.cleaned_data['usuario']
				clave = loginf.cleaned_data["contraseña"]
				user = authenticate(username = usuario, password = clave)
				try:
					user_temp = User.objects.get(username = usuario)
					user_id_temp = user_temp.id
					per_id= Persona.objects.get(usuario_id = user_id_temp)
					per_temp = per_id.id
					rol_temp = Rol_persona.objects.get(persona_id = per_temp)
					comprobacion_rol = str(rol_temp.rol)

					if comprobacion_rol == "ADMINISTRADOR" or comprobacion_rol == "VIGILANTE":
						if user is not None and user.is_active:
							login(request, user)
							usuarioActivo = True
							return redirect('url_index')
						else:
							msj = "usuario o clave incorrecto"
					else:
						info_enviada = True
						loginf = login_form()
						return render(request, "login.html", locals())
				except:
					msj = "El Usuario no existe"
		loginf = login_form()
		return render(request, "login.html", locals())
#=================================================#


#==================== LOGOUT =====================#
def view_logout(request):
	logout(request)
	return redirect('url_index')
#=================================================#

#============== LOGIN SUPERUSER ==================#
def view_login_superuser(request):
	error = "NO ES SUPERUSUARIO"
	superusuario = False
	if request.method == "POST":
		loginf = login_form(request.POST)
		if loginf.is_valid():
			usuario = loginf.cleaned_data['usuario']
			clave = loginf.cleaned_data["contraseña"]
			user = authenticate(username = usuario, password = clave)
			try:
				uSer = User.objects.get(username = usuario)
				superuser = uSer.is_superuser
				if superuser == True:
					superusuario = True
					try:
						if user is not None and user.is_active:
							login(request, user)
							return redirect('url_agregar_admin')
						else:
							msj = "usuario o clave incorrecto"
					except:
						msj = "El Usuario no existe"
				else:
					msj = 'No eres Super Usuario'
			except:
				msj = "El Usuario no existe"
	loginf = login_form()
	return render(request, "login_superadmin.html", locals())
#=================================================#


#==================GENERAR=REPORTE===================#
def reporte_ficha(request):
	fichas = Ficha.objects.all()
	fechaToday = date.today();

	if request.method == "POST":
		option_selected_numFicha = request.POST.get('numeroFicha', None)
		if option_selected_numFicha:
			queryFicha = Ficha.objects.filter(numeroFicha__contains = option_selected_numFicha)
			if queryFicha.exists():
				txtFicha = "Número de ficha"
				txtNombre = "Nombre del programa"
				ver = "Ver más"
			else:
				msjError = "Ficha "+option_selected_numFicha+" no encontrada"
		else:
			msjError = "Campo vacio"
	return render(request, "reporte_ficha.html", locals())
#===================================================#
#====================VIEW=FICHA=====================#
def view_ficha(request, id_ficha):
	ficha = Ficha.objects.get(id = id_ficha)
	persona = Persona_ficha.objects.filter(ficha__numeroFicha = ficha.numeroFicha)
	
	queryInstructor = Persona.objects.filter(id__in = [i.persona.id for i in Rol_persona.objects.filter(rol__rol="INSTRUCTOR")])

	permisoInstructor = []

	for queryI in queryInstructor:
		queryPermisoInstructor = Persona.objects.filter(id__in =[i.persona.id for i in Permiso_persona.objects.filter(persona__nombres = queryI)])
		if queryPermisoInstructor.exists():
			print("Permiso si existe de ",queryPermisoInstructor)
		else:
			print("No existe permiso de ",queryPermisoInstructor)
		permisoInstructor.append(queryPermisoInstructor)
	return render(request, 'view_ficha.html', locals())

#=================REPORTE=APRENDIZ===================#
def reporte_aprendiz(request):
	fichas = Ficha.objects.all()
	fechaToday = date.today();

	if request.method == "POST":

		option_selected_nombreAprendiz = request.POST.get('campoAprendiz', None)
		
		option_selected_numFicha = request.POST.get('numeroFicha', None)

		if option_selected_nombreAprendiz and not option_selected_numFicha:
			queryDocumentos = Persona_ficha.objects.filter(persona__documentoIdentidad__contains = option_selected_nombreAprendiz) 
			if queryDocumentos.exists():
				if queryDocumentos:
					txtDocumento = "Número de documento" 
					txtNombre = "Nombre"
					txtFicha = "Ficha"
					txtPerfil = "Ver perfil"
				else:
					msjError = "El número de documento "+option_selected_nombreAprendiz+" no se encuentra"
			else:
				queryNombres = Persona_ficha.objects.filter(persona__nombres__contains = option_selected_nombreAprendiz)
				if queryNombres.exists():
					txtDocumento = "Número de documento"
					txtNombre = "Nombre"
					txtFicha = "Ficha"
					txtPerfil = "Ver perfil"
				else:
					msjError = "El  aprendiz no se encuentra"
		elif option_selected_numFicha and not option_selected_nombreAprendiz:
			queryFicha = Persona_ficha.objects.filter(ficha__numeroFicha = option_selected_numFicha)
			if queryFicha.exists():
				txtDocumento = "Número de documento"
				txtNombre = "Nombre"
				txtFicha = "Ficha"
				txtPerfil = "Ver perfil"
			else:
				msjError = "Ficha vacía"
		else:
			queryDocumentoFicha = Persona_ficha.objects.filter(persona__documentoIdentidad__contains = option_selected_nombreAprendiz, ficha__numeroFicha = option_selected_numFicha)
			if queryDocumentoFicha:
				if queryDocumentoFicha.exists():
					txtDocumento = "Número de documento"
					txtNombre = "Nombre"
					txtFicha = "Ficha"
					txtPerfil = "Ver perfil"
				else:
					msjError = "El número de documento "+option_selected_nombreAprendiz+" no existe en la ficha: "+option_selected_numFicha
			else:
				queryNombreFicha = Persona_ficha.objects.filter(persona__nombres__contains = option_selected_nombreAprendiz, ficha__numeroFicha = option_selected_numFicha)
				if queryNombreFicha.exists():
					if queryNombreFicha:
						txtDocumento = "Número de documento"
						txtNombre = "Nombre"
						txtFicha = "Ficha"
						txtPerfil = "Ver perfil"
						btnPerfil = "Ver perfil"
				else:
					msjError = "El aprendiz no se encuentra en: "+option_selected_numFicha


	return render(request, "reporte_aprendiz.html", locals())

def view_perfil(request, id_persona):
	id = Persona.objects.get(id=id_persona)
	queryPermisos = Permiso_persona.objects.filter(persona__id = id_persona)
	try:
		queryRol = Rol_persona.objects.get(persona__id = id_persona)
	except:
		msjError = "No asignado"
	return render(request, "perfil.html", locals())

#==========================ESPRIN 3==========================================================================================================

#==================pETICIONES DE SALIDAS=======================#
def view_peticiones(request):

	peticion = Permiso_persona.objects.all()
	per_fic = Persona_ficha.objects.filter(persona__in=[i.persona.id for i in peticion])
	
	
	lis_pet = sorted(chain(peticion,per_fic),key=attrgetter('persona.id'))
	print(lis_pet)


	#query = Persona_ficha.objects.exclude(id__in=[i.persona.id for i in peticion])
	#print('-----------------------------------------------------------------------')
	#print(query)

	return render(request, "mostrar_peticiones.html", locals())

def view_ver_peticion(request, id_pet):

	permiso = Permiso_persona.objects.get(id=id_pet)
	ficha = Persona_ficha.objects.get(persona=permiso.persona.id)

	return render(request, "ver_peticion.html", locals())

def view_aprobar_peticion(request, id_pet):

	print('>>>>>>>>>APROBADO',id_pet)

	return redirect("url_peticiones")

def view_rechazar_peticion(request, id_pet):

	print('>>>>>>>>>No APROBADO',id_pet)

	return redirect("url_peticiones")